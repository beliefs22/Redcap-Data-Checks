import datetime
import openpyxl
from openpyxl.styles import PatternFill, colors

wb = openpyxl.load_workbook('Active_ED_Visit_Data_Check_10_6_16_testing.xlsx')
sheet = wb.active

new_fill = PatternFill('solid', fgColor=colors.RED)

def remove_num(num):
    for item in num[:]:
        if item.isdigit():
            num = num.replace(item,"")
    return num + "1"

def is_blank(cell):
    if cell.value == None:
        cell.fill = new_fill
        return "Value: %s, RedCap Label: %s" % \
               (str(cell.value), str(sheet[remove_num(cell.coordinate)].value))
    else:
        return ""

def yes_no_999(cell):
    if cell.value == "Yes" or cell.value =="No" or cell.value==999:
        return ""
    else:
        cell.fill = new_fill
        return "Value: %s, RedCap Label: %s" %\
               (str(cell.value), str(sheet[remove_num(cell.coordinate)].value))
def is_date(cell):
    if type(cell.value) != datetime.datetime:
        cell.fill = new_fill
        return "Value: %s, REDCap Label: %s" % \
               (str(cell.value), str(sheet[remove_num(cell.coordinate)].value))
    if cell.value.month not in range(1,13) or \
       cell.value.day not in range(1,32) or \
       cell.value.year not in range(2015,2017):
        cell.fill = new_fill
        return "Value: %s, REDCap Label: %s" % \
               (str(cell.value), str(sheet[remove_num(cell.coordinate)].value))    
    return ""

def is_time(cell):
    if type(cell.value) != datetime.time:
        cell.fill = new_fill
        return "Value: %s, REDCap Label: %s" % \
               (str(cell.value), str(sheet[remove_num(cell.coordinate)].value))    
    if cell.value.hour not in range(0, 25) or \
       cell.value.minute not in range(0, 60):
        cell.fill = new_fill
        return "Value: %s, REDCap Label: %s" % \
               (str(cell.value), str(sheet[remove_num(cell.coordinate)].value))
    
    return ""

def valid_temp(cell):
    if cell.value > 34.7 or cell.value < 40.8:
        return ""
    else:
        cell.fill = new_fill
        return "Value: %s, RedCap Label: %s" % \
               (str(cell.value), str(sheet[remove_num(cell.coordinate)].value))

def valid_pulse(cell):
    if cell.value > 50 or cell.value < 200:
        return ""
    else:
        cell.fill = new_fill
        return "Value: %s, RedCap Label: %s" % \
               (str(cell.value), str(sheet[remove_num(cell.coordinate)].value))

def valid_resp(cell):
    if cell.value > 10 or cell.value < 45:
        return ""
    else:
        cell.fill = new_fill
        return "Value: %s, RedCap Label: %s" % \
               (str(cell.value), str(sheet[remove_num(cell.coordinate)].value))

def valid_systolic(cell):
    if cell.value > 60 or cell.value < 240:
        return ""
    else:
        cell.fill = new_fill
        return "Value: %s, RedCap Label: %s" % \
               (str(cell.value), str(sheet[remove_num(cell.coordinate)].value))

def valid_oxygen(cell):
    if cell.value > 70 or cell.value < 100:
        return ""
    else:
        cell.fill = new_fill
        return "Value: %s, RedCap Label: %s" % \
               (str(cell.value), str(sheet[remove_num(cell.coordinate)].value))

def yes_no(cell):
    if cell.value == "Yes" or cell.value == "No":
        return ""
    else:
        cell.fill = new_fill
        return "Value: %s, RedCap Label: %s" % \
               (str(cell.value), str(sheet[remove_num(cell.coordinate)].value))

def yes_no_unknown(cell):
    if cell.value == "Yes" or cell.value == "No" or cell.value == "Unknown":
        return ""
    else:
        cell.fill = new_fill
        return "Value: %s, RedCap Label: %s" % \
               (str(cell.value), str(sheet[remove_num(cell.coordinate)].value))

def valid_ph(cell):
    if cell.value > 4 or cell.value < 10:
        return ""
    else:
        cell.fill = new_fill
        return "Value: %s, RedCap Label: %s" % \
               (str(cell.value), str(sheet[remove_num(cell.coordinate)].value))

def valid_bun(cell):
    if cell.value > 3 or cell.value < 30 or cell.value == 999:
        return ""
    else:
        cell.fill = new_fill
        return "Value: %s, RedCap Label: %s" % \
               (str(cell.value), str(sheet[remove_num(cell.coordinate)].value))

def valid_sodium(cell):
    if cell.value > 120 or cell.value < 147 or cell.value == 999:
        return ""
    else:
        cell.fill = new_fill
        return "Value: %s, RedCap Label: %s" % \
               (str(cell.value), str(sheet[remove_num(cell.coordinate)].value))

def valid_glucose(cell):
    if cell.value >50 or cell.value < 600 or cell.value == 999:
        return ""
    else:
        cell.fill = new_fill
        return "Value: %s, RedCap Label: %s" % \
               (str(cell.value), str(sheet[remove_num(cell.coordinate)].value))

def valid_hematocrit(cell):
    if cell.value > 15 or cell.value < 70:
        return ""
    else:
        cell.fill = new_fill
        return "Value: %s, RedCap Label: %s" % \
               (str(cell.value), str(sheet[remove_num(cell.coordinate)].value))

def valid_result(cell):
    if cell.value == "Negative" or cell.value == "Positive" or\
       cell.value == "Other":
        return ""
    else:
        cell.fill = new_fill
        return "Value: %s, RedCap Label: %s" % \
               (str(cell.value), str(sheet[remove_num(cell.coordinate)].value))

