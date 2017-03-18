import openpyxl
import collections
import validators_ed as val


def get_fields():
    with open('Master_Header_File_14_0076.xlsx', 'rb') as master_header_file:
        all_headers = openpyxl.load_workbook(master_header_file)

        # Get sheets for each file type
        eligibility_sheet = all_headers.get_sheet_by_name('Eligibility')
        demographic_sheet = all_headers.get_sheet_by_name('Demographic')
        symptom_sheet = all_headers.get_sheet_by_name('Symptoms')
        medical_sheet = all_headers.get_sheet_by_name('Medical')
        enrollment_specimen_sheet = all_headers.get_sheet_by_name('Enrollment_Specimen')
        follow_up_assessment_sheet = all_headers.get_sheet_by_name('Follow_Up_Assessment')
        follow_up_specimen_sheet = all_headers.get_sheet_by_name('Follow_Up_Specimen')
        ed_visit_sheet = all_headers.get_sheet_by_name('ED_Visits')
        ip_visit_sheet = all_headers.get_sheet_by_name('IP_Visits')

        all_sheets = {'eligibility': eligibility_sheet, 'demographics': demographic_sheet, 'symptoms': symptom_sheet,
                      'medical': medical_sheet, 'enrollment_specimen': enrollment_specimen_sheet,
                      'follow_up_assessment': follow_up_assessment_sheet,
                      'follow_up_specimen': follow_up_specimen_sheet, 'ed_visit': ed_visit_sheet,
                      'ip_visit': ip_visit_sheet}
        headers_and_fields = {}
        for sheet_name, sheet in all_sheets.iteritems():
            data_in_sheet = sheet.iter_rows()
            # first that correspond to the fields and data
            headers = [cell.value
                       for cell in data_in_sheet.next()
                       ]
            headers = headers[1:]

            fields = [[cell.value for cell in row
                       if cell.value]
                      for row in data_in_sheet
                      ]
            headers_and_fields[sheet_name] = {'headers': headers,
                                              'fields': fields}
        return headers_and_fields


def test_by_type(cell, test_type):
    test = {'date': val.is_date, 'time': val.is_time, 'pulse': val.valid_pulse, 'temp': val.valid_temp,
            'resp': val.valid_resp, 'systolic': val.valid_systolic, 'sodium': val.valid_sodium,
            'glucose': val.valid_glucose, 'hematocrit': val.valid_hematocrit, 'oxygen': val.valid_oxygen,
            'bun': val.valid_bun, 'blank': val.is_blank, 'ph': val.valid_ph
            }

    return test[test_type](cell)


def simple_check(field, tuple, statement):

    simple_field_names = field[1:]
    for field_name in simple_field_names:
        eval("val.is_blank(" + statement % (field_name,) + ")")


def complex_without_number_check(field, tuple, statement):
    start_check = field[1]
    complex_field_names = field[2:]
    # check start condition
    if eval(statement % start_check) == True:
        for field_name in complex_field_names:
            eval("val.is_blank(" + statement % (field_name,) + ")")


def complex_with_number_check(field, tuple, statement):
    start_check = field[1]
    number_to_check = eval(statement % field[2])
    complex_field_names = field[3:]
    # check start condition
    if eval(statement % start_check) == True:
        for i in range(1, number_to_check + 1):
            for field_name in complex_field_names:
                eval("val.is_blank(" + statement % (field_name % i,) + ")")


def simple_visit_check(field, tuple, statement, visit_num):
    if visit_num:
        simple_field_names = field[1:]
        for field_name in simple_field_names:
            eval("val.is_blank(" + statement % (field_name % visit_num,) + ")")


def complex_visit_with_number_check(field, tuple, statement, visit_num):
    if visit_num:
        start_check = field[1] % visit_num
        number_to_check = eval(statement % field[2] % visit_num)
        if number_to_check == "more than three":
            number_to_check = 3
        complex_field_names = field[3:]
        # check start condition
        if number_to_check >= 1:
            if eval(statement % start_check) == True:
                for i in range(1, number_to_check + 1):
                    for field_name in complex_field_names:
                        eval("val.is_blank(" + statement % (field_name % (visit_num, i) + ")"))


def complex_visit_without_number_check(field, tuple, statement, visit_num):
    if visit_num:
        start_check = field[1] % visit_num
        complex_field_names = field[2:]
        # check start condition
        if eval(statement % start_check) == True:
            for field_name in complex_field_names:
                eval("val.is_blank(" + statement % (field_name % visit_num,) + ")")


def complex_visit_with_complex_number(field, tuple, statement, visit_num):
    if visit_num:
        number_to_check = eval(statement % (field[1] % visit_num))
        if number_to_check == 'more than three':
            number_to_check = 3
        if number_to_check >= 1:
            for i in range(1, number_to_check + 1):
                start_check = field[2] % (visit_num, i)
                complex_field_names = field[3:]
                # check start condition
                if eval(statement % start_check) == True:
                    for field_name in complex_field_names:
                        eval("val.is_blank(" + statement % (field_name % (visit_num, i) + ")"))


def generic_check(check_type, file_to_check, headers, fields):
    print "startin %s check" % check_type
    with open(file_to_check, 'rb') as datafile:
        data_workbook = openpyxl.load_workbook(datafile)
        data_worksheet = data_workbook.active
        data_headers = headers
        data_fields = fields
        data_tuple = collections.namedtuple(check_type, field_names=headers)
        data = data_worksheet.iter_rows()
        # skip human readable headers in data file
        data.next()
        data_to_check = []
        for row in data:
            one_subjects_data = data_tuple(*[cell for cell in row
                                             ])
            data_to_check.append(one_subjects_data)

        statement = "tuple.%s"
        for tuple in data_to_check:
            for field in data_fields:
                field_type = field[0]
                if field_type == 'simple:':
                    simple_check(field, tuple, statement)
                if field_type == 'complex with number':
                    complex_with_number_check(field, tuple, statement)
                if field_type == 'complex without number':
                    complex_without_number_check(field, tuple, statement)
                if field_type == 'simple visit without number':
                    visit_num = tuple.visit_num.value
                    simple_visit_check(field, tuple, statement, visit_num)
                if field_type == 'complex visit with number':
                    visit_num = tuple.visit_num.value
                    complex_visit_with_number_check(field, tuple, statement, visit_num)
                if field_type == 'complex visit without number':
                    visit_num = tuple.visit_num.value
                    complex_visit_without_number_check(field, tuple, statement, visit_num)
                if field_type == 'complex visit with complex number':
                    visit_num = tuple.visit_num.value
                    complex_visit_with_complex_number(field, tuple, statement, visit_num)

        data_workbook.save(check_type + "_with_highlight_errors.xlsx")


def main():

    all_fields = get_fields()
    files_to_test = {'eligibility' : [all_fields['eligibility'],'14_0076_Eligibility.xlsx'],
                     'demographics' : [all_fields['demographics'], '14_0076_Demographic.xlsx'],
                     'symptoms' : [all_fields['symptoms'],'14_0076_Symptoms.xlsx'],
                     'medical' : [all_fields['medical'], '14_0076_Medical.xlsx'],
                     'enrollment_specimen': [all_fields['enrollment_specimen'],'14_0076_Enrollment_Specimen.xlsx'],
                     'follow_up_assessment' : [all_fields['follow_up_assessment'],'14_0076_Follow_Up_Assessment.xlsx'],
                     'follow_up_specimen': [all_fields['follow_up_specimen'],'14_0076_Follow_Up_Specimen.xlsx'],
                     'ed_visit' : [all_fields['ed_visit'],'14_0076_ED_Visit.xlsx'],
                     'ip_visit': [all_fields['ip_visit'],'14_0076_IP_Visit.xlsx']}

    for check_type, file_info in files_to_test.iteritems():
        check_type = check_type
        file_to_check = file_info[1]
        headers = file_info[0]['headers']
        fields = file_info[0]['fields']
        generic_check(check_type,file_to_check,headers,fields)



if __name__ == "__main__":
    main()