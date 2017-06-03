import openpyxl
import collections
import csv
import validators_ed as val

def get_fields():
    with open('Master_Header_File_14_0076_w_form_names.xlsx', 'rb') as master_header_file:
        all_headers = openpyxl.load_workbook(master_header_file)

        #Get sheets for each file type
        eligibility_sheet = all_headers.get_sheet_by_name('Eligibility')
        demographic_sheet = all_headers.get_sheet_by_name('Demographic')
        symptom_sheet = all_headers.get_sheet_by_name('Symptoms')
        medical_sheet = all_headers.get_sheet_by_name('Medical')
        enrollment_specimen_sheet = all_headers.get_sheet_by_name('Enrollment_Specimen')
        follow_up_assessment_sheet = all_headers.get_sheet_by_name('Follow_Up_Assessment')
        follow_up_specimen_sheet = all_headers.get_sheet_by_name('Follow_Up_Specimen')
        ed_visit_sheet = all_headers.get_sheet_by_name('ED_Visits')
        ip_visit_sheet = all_headers.get_sheet_by_name('IP_Visits')

        all_sheets = {'eligibility': eligibility_sheet, 'demographic': demographic_sheet, 'symptoms': symptom_sheet,
                      'medical': medical_sheet, 'enrollment_specimen': enrollment_specimen_sheet,
                      'follow_up_assessment' : follow_up_assessment_sheet,
                      'follow_up_specimen': follow_up_specimen_sheet, 'ed_visit': ed_visit_sheet,
                      'ip_visit':ip_visit_sheet}
        headers_fields_forms = {}
        for sheet_name, sheet in all_sheets.iteritems():
            data_in_sheet = sheet.iter_rows()
            #first row contains header information
            headers = [unicode(cell.value)
                       for cell in data_in_sheet.next()
                       ]
            # machine_headers = [unicode(cell.value)
            #            for cell in data_in_sheet.next()
            #            ]
            # machine_headers = machine_headers[1:]
            # human_headers = [unicode(cell.value)
            #                  for cell in data_in_sheet.next()
            #                  ]
            # human_headers = human_headers[1:]
            # headers = {machine_header:human_header
            #            for machine_header,human_header in zip(machine_headers,human_headers)}
            #Secon row contains information linking paper forms to headers
            form_info = [unicode(cell.value)
                         for cell in data_in_sheet.next()
                         ]
            form_info = form_info[1:]
            forms = {header:form
                     for header, form in zip(headers,form_info)
                     }
            #Remaining rows contain fields to validate
            fields = [[cell.value for cell in row
                       if cell.value]
                      for row in data_in_sheet
                      ]
            headers_fields_forms[sheet_name] = {'headers': headers,
                                                'fields': fields,
                                                'forms': forms
                                                }
    return headers_fields_forms

def test_by_type(cell, test_type):
    test = {'date': val.is_date, 'time': val.is_time, 'pulse': val.valid_pulse,'temp': val.valid_temp,
            'resp': val.valid_resp, 'systolic': val.valid_systolic, 'sodium': val.valid_sodium,
            'glucose': val.valid_glucose, 'hematocrit': val.valid_hematocrit, 'oxygen': val.valid_oxygen,
            'bun': val.valid_bun, 'blank': val.is_blank, 'ph': val.valid_ph
            }

    return test[test_type](cell)

def simple_check(field, tuple, statement):
    errors = []
    tuple = tuple
    simple_field_names = field[1:]
    for field_name in simple_field_names:
        error_found = eval(statement % (field_name,) + ".value") == None
        if error_found:
            errors.append(field_name)
    return errors


def complex_without_number_check(field, tuple, statement):
    errors = []
    tuple = tuple
    start_check = field[1]
    complex_field_names = field[2:]
    # check start condition
    if eval(statement % start_check) == True:
        for field_name in complex_field_names:
            error_found = eval(statement % (field_name,) + ".value") == None
            if error_found:
                errors.append(field_name)
    return errors

def complex_with_number_check(field, tuple, statement):
    errors = []
    tuple = tuple
    start_check = field[1]
    number_to_check = eval(statement % field[2])
    complex_field_names = field[3:]
    # check start condition
    if number_to_check:
        if eval(statement % start_check) == True:
            for i in range(1, number_to_check + 1):
                for field_name in complex_field_names:
                    error_found = eval(statement % (field_name % i,) + ".value") == None
                    if error_found:
                        errors.append(field_name % i)
    return errors

def simple_visit_check(field, tuple, statement, visit_num):
    errors = []
    if visit_num:
        tuple = tuple
        simple_field_names = field[1:]
        for field_name in simple_field_names:
            error_found = eval(statement % (field_name % visit_num,) + ".value") == None
            if error_found:
                errors.append(field_name % visit_num)
    return errors

def complex_visit_with_number_check(field, tuple, statement, visit_num):
    errors = []
    if visit_num:
        tuple = tuple
        start_check = field[1] % visit_num
        number_to_check = eval(statement % field[2] % visit_num)
        if number_to_check:
            if number_to_check == "more than three":
                number_to_check = 3
            complex_field_names = field[3:]
            # check start condition
            if number_to_check >= 1:
                if eval(statement % start_check) == True:
                    for i in range(1, number_to_check + 1):
                        for field_name in complex_field_names:
                            error_found = eval(statement % (field_name % (visit_num, i)) + ".value") == None
                            if error_found:
                                errors.append(field_name % (visit_num, i))
    return errors


def complex_visit_without_number_check(field, tuple, statement, visit_num):
    errors = []
    if visit_num:
        tuple = tuple
        start_check = field[1] % visit_num
        complex_field_names = field[2:]
        # check start condition
        if eval(statement % start_check) == True:
            for field_name in complex_field_names:
                error_found = eval(statement % (field_name % visit_num,) + ".value") == None
                if error_found:
                    errors.append(field_name % visit_num)
    return errors

def complex_visit_with_complex_number(field, tuple, statement, visit_num):
    errors = []
    if visit_num:
        tuple = tuple
        number_to_check = eval(statement % (field[1] % visit_num))
        if number_to_check:
            if number_to_check == 'more than three':
                number_to_check = 3
            if number_to_check >= 1:
                for i in range(1, number_to_check + 1):
                    start_check = field[2] % (visit_num, i)
                    complex_field_names = field[3:]
                    # check start condition
                    if eval(statement % start_check) == True:
                        for field_name in complex_field_names:
                            error_found = eval(statement % (field_name % (visit_num, i)) + ".value") == None
                            if error_found:
                                errors.append(field_name % (visit_num, i))
    return errors


def generic_check(check_type, file_to_check, headers, fields, forms):
    print "startin %s check" % check_type
    with open(file_to_check,'rb') as datafile:
        data_workbook = openpyxl.load_workbook(datafile)
        data_worksheet = data_workbook.active
        data_fields = fields
        data_tuple = collections.namedtuple(check_type, field_names=headers)
        data = data_worksheet.iter_rows()
        #skip human readable headers in data file
        data.next()
        data_to_check = []
        for row in data:
            for_tuple = [cell for cell in row]
            one_subjects_data = data_tuple(*for_tuple)
            data_to_check.append(one_subjects_data)

        statement = "tuple.%s"
        all_found_errors = list()
        for tuple in data_to_check:
            errors = []
            # print tuple.id.value
            for field in data_fields:
                field_type = field[0]
                if field_type == 'simple':
                    error_found = simple_check(field, tuple, statement)
                    if error_found:
                        for error in error_found:
                            redcap_label = error
                            form_name = forms['id']
                            question_info = forms[error]
                            actual_value = eval(statement % error + ".value")
                            errors.append([tuple.id.value , form_name, question_info, redcap_label, actual_value])
                if field_type == 'complex with number':
                    error_found = complex_with_number_check(field, tuple, statement)
                    if error_found:
                        for error in error_found:
                            redcap_label = error
                            form_name = forms['id']
                            question_info = forms[error]
                            actual_value = eval(statement % error + ".value")
                            errors.append([tuple.id.value , form_name, question_info, redcap_label, actual_value])
                if field_type == 'complex without number':
                    error_found = complex_without_number_check(field, tuple, statement)
                    if error_found:
                        for error in error_found:
                            redcap_label = error
                            form_name = forms['id']
                            question_info = forms[error]
                            actual_value = eval(statement % error + ".value")
                            errors.append([tuple.id.value , form_name, question_info, redcap_label, actual_value])
                if field_type == 'simple visit without number':
                    visit_num = tuple.visit_num.value
                    error_found = simple_visit_check(field, tuple, statement, visit_num)
                    if error_found:
                        for error in error_found:
                            redcap_label = error
                            form_name = forms['id']
                            question_info = forms[error]
                            actual_value = eval(statement % error + ".value")
                            errors.append([tuple.id.value , form_name, question_info, redcap_label, actual_value])
                if field_type == 'complex visit with number':
                    visit_num = tuple.visit_num.value
                    error_found = complex_visit_with_number_check(field, tuple, statement, visit_num)
                    if error_found:
                        for error in error_found:
                            redcap_label = error
                            form_name = forms['id']
                            question_info = forms[error]
                            actual_value = eval(statement % error + ".value")
                            errors.append([tuple.id.value , form_name, question_info, redcap_label, actual_value])
                if field_type == 'complex visit without number':
                    visit_num = tuple.visit_num.value
                    error_found = complex_visit_without_number_check(field, tuple, statement, visit_num)
                    if error_found:
                        for error in error_found:
                            redcap_label = error
                            form_name = forms['id']
                            question_info = forms[error]
                            actual_value = eval(statement % error + ".value")
                            errors.append([tuple.id.value , form_name, question_info, redcap_label, actual_value])
                if field_type == 'complex visit with complex number':
                    visit_num = tuple.visit_num.value
                    error_found = complex_visit_with_complex_number(field, tuple, statement, visit_num)
                    if error_found:
                        for error in error_found:
                            redcap_label = error
                            form_name = forms['id']
                            question_info = forms[error]
                            actual_value = eval(statement % error + ".value")
                            errors.append([tuple.id.value , form_name, question_info, redcap_label, actual_value])
            if errors != []:
                all_found_errors.append(errors)

        data_workbook.save(check_type + "_with_highlight_errors.xlsx")
        for item in all_found_errors:
            for single_error in item:
                print single_error
def main():
    fields = get_fields()
    for check_type, headers_fields_forms in fields.iteritems():
        filename = "14_0076_{}.xlsx".format(check_type)
        print filename
        headers = headers_fields_forms['headers']
        fields = headers_fields_forms['fields']
        forms = headers_fields_forms['forms']
        # print fields
        generic_check(check_type,filename,headers,fields, forms)

if __name__ == '__main__':
    main()