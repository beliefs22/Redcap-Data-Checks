import collections
import csv
import openpyxl
import checks_ed_old
import checks_ip_updated
from openpyxl.styles import PatternFill, colors

new_fill = PatternFill('solid', fgColor=colors.RED)
#first file has the machine readable headers
#second file contains human readable headers and the data to check
with open('CEIRSActiveSurveilla_DATA_2016-12-18_1135.csv', 'rb') as myfile,\
    open('CEIRSActiveSurveilla_DATA_LABELS_2016-12-18_1137.xlsx', 'rb') as testfile, \
    open('output_ed.txt', 'w') as outfile:
    #make csv reader object for out of machine header file, to get headers, we use this to match machine
    #headers to human readable headers
    csvreader = csv.DictReader(myfile)
    headers = csvreader.fieldnames
    subject_data = openpyxl.load_workbook(testfile)
    #the excel sheet that containts the data to check should be the active sheet [should we call it by name?]
    data_sheet = subject_data.active
    #named tuple which will have a field for each header. THe field names are the machine readable headers. We use the
    #machine readable headers because they are unique where the human readable headers repeat
    ed_tuple = collections.namedtuple('ED', field_names=headers)
    rows = data_sheet.iter_rows()
    readable_headers = rows.next()
    mytuples = []
    #create a named tuple for each data entry in the excel file
    for row in rows:
        one_visit = ed_tuple(*[cell
                               for cell in row
                               ])
        mytuples.append(one_visit)
    results = []
    #simple fields are fields that we just need to check if they are blank or valid types i.e. enrollment date, gender
    #complex fields are fields that are interdepend for example if the subject had a flu test, then we need to check
    # test name, date, result, etcs..
    complex, simple = checks_ed_old.field_names()
    for atuple in mytuples:
        #we only  check rows that are marked completed and ready for QC
        if atuple.form_10a_ed_chart_review_active_surveillance_complete.value != 'Complete':
            continue
        answer = checks_ed_old.simple_blank_check(atuple, simple)
        if answer:
            results.append((atuple.id.value, answer))
        for field_names in complex:
            print "looking for fields", field_names
            answer = checks_ed_old.complex_check(atuple, complex[field_names])
            if answer:
                results.append((atuple.id.value, answer))

    final_result = collections.defaultdict(list)
    for item in results:
        final_result[item[0]] += item[1]
    print "writing output file"
    for key, values in final_result.iteritems():
        outfile.write("Subject {} Errors:\n".format(key))
        for index, value in enumerate(values):
            outfile.write("\t\t\t\t\t\t\t{}. {}\n".format(index + 1, value))
        outfile.write("\n------------------------------------------------------------------------------------------\n")

    print "finished output file"
    print "saving output excel"
    subject_data.save('ed_text.xlsx')
    print "finished"

with open('CEIRSActiveSurveilla_DATA_2016-12-21_1806.csv', 'rb') as myfile,\
    open('CEIRSActiveSurveilla_DATA_LABELS_2016-12-21_1806.xlsx', 'rb') as testfile, \
    open('output_ip.txt', 'w') as outfile:
    csvreader = csv.DictReader(myfile)
    headers = csvreader.fieldnames
    subject_data = openpyxl.load_workbook(testfile)
    data_sheet = subject_data.active

    ip_tuple = collections.namedtuple('IP', field_names=headers)
    rows = data_sheet.iter_rows()
    rows.next()
    mytuples = []
    for row in rows:
        one_visit = ip_tuple(*[cell
                               for cell in row
                               ])
        mytuples.append(one_visit)
    results = []
    complex, simple = checks_ip_updated.field_names()
    for atuple in mytuples:
        if atuple.form_11a_chart_review_inpatient_hospitalization_complete.value != 'Complete':
            continue
        for field_names in complex:
            print "looking for fields", field_names
            answer = checks_ip_updated.complex_check(atuple, complex[field_names])
            if answer:
                results.append((atuple.id.value, answer))

    final_result = collections.defaultdict(list)
    for item in results:
        final_result[item[0]] += item[1]
    print "writing IP output file"
    for key, values in final_result.iteritems():
        outfile.write("Subject {} Errors:\n".format(key))
        for index, value in enumerate(values):
            outfile.write("\t\t\t\t\t\t\t{}. {}\n".format(index + 1, value))
        outfile.write("\n------------------------------------------------------------------------------------------\n")

    print "finished output file"
    print "saving output excel"
    subject_data.save('IP_text.xlsx')
    print "finished"
