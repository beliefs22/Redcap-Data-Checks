import openpyxl
import collections
import csv
# import generic_validators as val

def get_fields(protocol):
    with open('{}_raw_Data.xlsx'.format(protocol), 'rb') as raw_header_file,open('{}_labeled_data.xlsx'.format(protocol), 'rb') as labeled_header_file:
        raw_header_book = openpyxl.load_workbook(raw_header_file)
        labeled_header_book = openpyxl.load_workbook(labeled_header_file)
        raw_header_sheet = raw_header_book.active
        raw_header_data = raw_header_sheet.iter_rows()
        labeled_header_sheet = labeled_header_book.active
        labeled_header_data = labeled_header_sheet.iter_rows()
        machine_headers = [unicode(cell.value)
                           for cell in raw_header_data.next()
                           ]
        #replace mispelled headers
        machine_headers[machine_headers.index('curmedsteriod1dose')] = 'curmedsteroid1dose'
        machine_headers = [header.replace('typingsp','typsp')
                           for header in machine_headers
                           ]
        human_headers = [unicode(cell.value)
                         for cell in labeled_header_data.next()
                         ]
        headers = collections.OrderedDict()
        for machine_header, human_header in zip(machine_headers, human_headers):
            headers[machine_header] = human_header
        with open('Master_Header_File_{}.xlsx'.format(protocol), 'rb') as master_header_file:
            all_headers = openpyxl.load_workbook(master_header_file)
            sheet_names = all_headers.get_sheet_names()
            all_sheets = {sheet_name.lower() : all_headers.get_sheet_by_name(sheet_name)
                          for sheet_name in sheet_names}
            headers_fields_forms = {}
            for sheet_name, sheet in all_sheets.iteritems():
                data_in_sheet = sheet.iter_rows()
                #Get machine and human headers from sheet
                form_machine_headers = [unicode(cell.value)
                                        for cell in data_in_sheet.next()
                                        ]
                form_machine_headers = form_machine_headers[1:]
                form_human_headers = [unicode(cell.value)
                                      for cell in data_in_sheet.next()
                                      ]
                form_human_headers = form_human_headers[1:]
                form_info = [unicode(cell.value)
                             for cell in data_in_sheet.next()
                             ]
                form_info = form_info[1:]
                forms = {header:form
                         for header, form in zip(form_machine_headers,form_info)
                         }
                fields = [[cell.value for cell in row
                           if cell.value]
                          for row in data_in_sheet
                          ]
                headers_fields_forms[sheet_name] = {'headers': headers,
                                                    'fields': fields,
                                                    'forms': forms
                                                    }
    return headers_fields_forms


def test_by_type(cell, test_type):
    test = {'date': val.is_date, 'time': val.is_time, 'pulse': val.valid_pulse,'temp': val.valid_temp,
            'resp': val.valid_resp, 'systolic': val.valid_systolic, 'sodium': val.valid_sodium,
            'glucose': val.valid_glucose, 'hematocrit': val.valid_hematocrit, 'oxygen': val.valid_oxygen,
            'bun': val.valid_bun, 'blank': val.is_blank, 'ph': val.valid_ph
            }

    return test[test_type](cell)

def simple_check(field, tuple, statement):
    errors = []
    tuple = tuple
    simple_field_names = field[1:]
    for field_name in simple_field_names:
        error_found = eval(statement % (field_name,))== ""
        if error_found:
            errors.append(field_name)
    return errors


def complex_without_number_check(field, tuple, statement):
    errors = []
    tuple = tuple
    start_check = field[1]
    complex_field_names = field[2:]
    # check start condition
    if eval(statement % start_check) == True:
        for field_name in complex_field_names:
            error_found = eval(statement % (field_name,))== ""
            if error_found:
                errors.append(field_name)
    return errors

def complex_with_number_check(field, tuple, statement):
    errors = []
    tuple = tuple
    start_check = field[1]
    number_to_check = eval(statement % field[2])
    complex_field_names = field[3:]
    # check start condition
    if number_to_check:
        if number_to_check == 'more than three':
            number_to_check = 3
        number_to_check = int(number_to_check)
        if eval(statement % start_check) == True:
            for i in range(1, number_to_check + 1):
                for field_name in complex_field_names:
                    error_found = eval(statement % (field_name % i,))== ""
                    if error_found:
                        errors.append(field_name % i)
    return errors

def simple_visit_check(field, tuple, statement, visit_num):
    errors = []
    if visit_num:
        visit_num = int(visit_num)
        if visit_num != 0:
            tuple = tuple
            simple_field_names = field[1:]
            for field_name in simple_field_names:
                error_found = eval(statement % (field_name % visit_num,))== ""
                if error_found:
                    errors.append(field_name % visit_num)
    return errors

def complex_visit_with_number_check(field, tuple, statement, visit_num):
    errors = []
    if visit_num:
        visit_num = int(visit_num)
        if visit_num != 0:
            tuple = tuple
            start_check = field[1] % visit_num
            number_to_check = eval(statement % field[2] % visit_num)
            if number_to_check:
                if number_to_check == "more than three":
                    number_to_check = 3
                else:
                    number_to_check = int(number_to_check)
                complex_field_names = field[3:]
                # check start condition
                if number_to_check >= 1:
                    if eval(statement % start_check) == True:
                        for i in range(1, number_to_check + 1):
                            for field_name in complex_field_names:
                                error_found = eval(statement % (field_name % (visit_num, i)))== ""
                                if error_found:
                                    errors.append(field_name % (visit_num, i))
    return errors


def complex_visit_without_number_check(field, tuple, statement, visit_num):
    errors = []
    if visit_num:
        visit_num = int(visit_num)
        if visit_num != 0:
            tuple = tuple
            start_check = field[1] % visit_num
            complex_field_names = field[2:]
            # check start condition
            if eval(statement % start_check) == True:
                for field_name in complex_field_names:
                    error_found = eval(statement % (field_name % visit_num,))== ""
                    if error_found:
                        errors.append(field_name % visit_num)
    return errors

def complex_visit_with_complex_number(field, tuple, statement, visit_num):
    errors = []
    if visit_num:
        visit_num = int(visit_num)
        if visit_num != 0:
            tuple = tuple
            number_to_check = eval(statement % (field[1] % visit_num))
            if number_to_check:
                if number_to_check == 'more than three':
                    number_to_check = 3
                else:
                    number_to_check = int(number_to_check)
                if number_to_check >= 1:
                    for i in range(1, number_to_check + 1):
                        start_check = field[2] % (visit_num, i)
                        complex_field_names = field[3:]
                        # check start condition
                        if eval(statement % start_check) == True:
                            for field_name in complex_field_names:
                                error_found = eval(statement % (field_name % (visit_num, i)))== ""
                                if error_found:
                                    errors.append(field_name % (visit_num, i))
    return errors

def complex_with_complex_number(field, tuple, statement):
    errors = []
    tuple = tuple
    number_to_check = eval(statement % (field[1]))
    if number_to_check:
        if number_to_check == 'more than three':
            number_to_check = 3
        else:
            number_to_check = int(number_to_check)
        if number_to_check >= 1:
            for i in range(1, number_to_check + 1):
                start_check = field[2] % i
                complex_field_names = field[3:]
                if eval(statement % start_check) == True:
                    for field_name in complex_field_names:
                        error_found = eval(statement % (field_name % i))== ""
                        if error_found:
                            errors.append(field_name % i)
    return errors


def generic_check(check_type, file_to_check, headers, fields, forms):
    print "startin %s check" % check_type
    all_found_errors = list()
    visit_number_locations = {'ed_visit': 'edptchart_visitnumber',
                              'ip_visit': 'inptchart_visitnumber'
    }
    with open(file_to_check,'rb') as datafile:
        data_reader = csv.reader(datafile)
        data_fields = fields
        data_tuple = collections.namedtuple(check_type, field_names=headers.keys())
        #skip human readable headers in data file
        csv_headers = data_reader.next()
        data_to_check = []
        for row in data_reader:
            if row:
                for_tuple = [cell for cell in row]
                one_subjects_data = data_tuple(*for_tuple)
                form_name = forms['id']
                complete = eval('one_subjects_data.{}'.format(form_name + "_complete"))
                if complete in ('Unverified', 'Complete'):
                    data_to_check.append(one_subjects_data)

        statement = "tuple.%s"
        for tuple in data_to_check:
            errors = []
            # print tuple.id.value
            for field in data_fields:
                error_message = field[-1]
                field = field[:len(field)-1]
                field_type = field[0]
                if field_type == 'simple':
                    error_found = simple_check(field, tuple, statement)
                    if error_found:
                        for error in error_found:
                            redcap_label = error
                            form_name = forms['id']
                            question_info = forms[error]
                            actual_value = eval(statement % error)
                            errors.append([tuple.id , form_name, question_info, redcap_label, actual_value, error_message])
                if field_type == 'complex with number':
                    error_found = complex_with_number_check(field, tuple, statement)
                    if error_found:
                        for error in error_found:
                            redcap_label = headers[error]
                            form_name = forms['id']
                            question_info = forms[error]
                            actual_value = eval(statement % error)
                            errors.append([tuple.id , form_name, question_info, redcap_label, actual_value, error_message])
                if field_type == 'complex with complex number':
                    error_found = complex_with_complex_number(field, tuple, statement)
                    if error_found:
                        for error in error_found:
                            redcap_label = headers[error]
                            form_name = forms['id']
                            question_info = forms[error]
                            actual_value = eval(statement % error)
                            errors.append([tuple.id , form_name, question_info, redcap_label, actual_value, error_message])
                if field_type == 'complex without number':
                    error_found = complex_without_number_check(field, tuple, statement)
                    if error_found:
                        for error in error_found:
                            redcap_label = headers[error]
                            form_name = forms['id']
                            question_info = forms[error]
                            actual_value = eval(statement % error)
                            errors.append([tuple.id , form_name, question_info, redcap_label, actual_value, error_message])
                if field_type == 'simple visit without number':
                    visit_num = eval("tuple.{}".format(visit_number_locations[check_type]))
                    if check_type == 'ip_visit':
                        if int(visit_num) > 3:
                            visit_num = 3
                    if check_type == 'ed_visit':
                        if int(visit_num) > 4:
                            visit_num = 4
                    error_found = simple_visit_check(field, tuple, statement, visit_num)
                    if error_found:
                        for error in error_found:
                            redcap_label = headers[error]
                            form_name = forms['id']
                            question_info = forms[error]
                            actual_value = eval(statement % error)
                            errors.append([tuple.id , form_name, question_info, redcap_label, actual_value, error_message])
                if field_type == 'complex visit with number':
                    visit_num = eval("tuple.{}".format(visit_number_locations[check_type]))
                    if check_type == 'ip_visit':
                        if int(visit_num) > 3:
                            visit_num = 3
                    if check_type == 'ed_visit':
                        if int(visit_num) > 4:
                            visit_num = 4
                    error_found = complex_visit_with_number_check(field, tuple, statement, visit_num)
                    if error_found:
                        for error in error_found:
                            redcap_label = headers[error]
                            form_name = forms['id']
                            question_info = forms[error]
                            actual_value = eval(statement % error)
                            errors.append([tuple.id , form_name, question_info, redcap_label, actual_value, error_message])
                if field_type == 'complex visit without number':
                    visit_num = eval("tuple.{}".format(visit_number_locations[check_type]))
                    if check_type == 'ip_visit':
                        if int(visit_num) > 3:
                            visit_num = 3
                    if check_type == 'ed_visit':
                        if int(visit_num) > 4:
                            visit_num = 4
                    error_found = complex_visit_without_number_check(field, tuple, statement, visit_num)
                    if error_found:
                        for error in error_found:
                            redcap_label = headers[error]
                            form_name = forms['id']
                            question_info = forms[error]
                            actual_value = eval(statement % error)
                            errors.append([tuple.id , form_name, question_info, redcap_label, actual_value, error_message])
                if field_type == 'complex visit with complex number':
                    visit_num = eval("tuple.{}".format(visit_number_locations[check_type]))
                    if check_type == 'ip_visit':
                        if int(visit_num) > 3:
                            visit_num = 3
                    if check_type == 'ed_visit':
                        if int(visit_num) > 4:
                            visit_num = 4
                    error_found = complex_visit_with_complex_number(field, tuple, statement, visit_num)
                    if error_found:
                        for error in error_found:
                            redcap_label = headers[error]
                            form_name = forms['id']
                            question_info = forms[error]
                            actual_value = eval(statement % error)
                            errors.append([tuple.id , form_name, question_info, redcap_label, actual_value, error_message])
            if errors != []:
                all_found_errors.append(errors)
    return all_found_errors

def main():
    protocols = ['14_0076', '15_0103']
    for protocol in protocols:
        fields = get_fields(protocol)
        all_found_errors = list()
        for check_type, headers_fields_forms in fields.iteritems():
            filename = "{}_labeled_data.csv".format(protocol)
            print filename
            headers = headers_fields_forms['headers']
            fields = headers_fields_forms['fields']
            forms = headers_fields_forms['forms']
            # print fields
            errors = generic_check(check_type,filename,headers,fields, forms)
            if errors:
                all_found_errors.append(errors)
        with open('{}_errors.csv'.format(protocol), 'wb') as error_file:
            csvwriter = csv.writer(error_file)
            csv_headers = ['Subject ID', 'Form Name', 'CRF Question Location', 'RedCap Label', 'Current Value', 'Details']
            csvwriter.writerow(csv_headers)
            for item in all_found_errors:
                for subjects_errors in item:
                    for single_error in subjects_errors:
                        single_error = [item.encode('ascii','ignore')
                                        for item in single_error]
                        csvwriter.writerow(single_error)

if __name__ == '__main__':
    main()