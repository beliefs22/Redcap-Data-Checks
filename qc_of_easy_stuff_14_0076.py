import openpyxl
import collections
import csv
import validators_ed as val

def get_fields():
    with open('Master_Header_File_14_0076.xlsx', 'rb') as master_header_file:
        all_headers = openpyxl.load_workbook(master_header_file)

        #Get sheets for each file type
        eligibility_sheet = all_headers.get_sheet_by_name('Eligibility')
        demographic_sheet = all_headers.get_sheet_by_name('Demographic')
        symptom_sheet = all_headers.get_sheet_by_name('Symptoms')
        medical_sheet = all_headers.get_sheet_by_name('Medical')
        enrollment_specimen_sheet = all_headers.get_sheet_by_name('Enrollment_Specimen')
        follow_up_assessment_sheet = all_headers.get_sheet_by_name('Follow_Up_Assessment')
        follow_up_specimen_sheet = all_headers.get_sheet_by_name('Follow_Up_Specimen')
        ed_visit_sheet = all_headers.get_sheet_by_name('ED_Visits')
        ip_visit_sheet = all_headers.get_sheet_by_name('IP_Visits')

        all_sheets = {'eligibility': eligibility_sheet, 'demographic': demographic_sheet, 'symptoms': symptom_sheet,
                      'medical': medical_sheet, 'enrollment_specimen': enrollment_specimen_sheet,
                      'follow_up_assessment' : follow_up_assessment_sheet,
                      'follow_up_specimen': follow_up_specimen_sheet, 'ed_visit': ed_visit_sheet,
                      'ip_visit':ip_visit_sheet}
        headers_and_fields = {}
        for sheet_name, sheet in all_sheets.iteritems():
            data_in_sheet = sheet.iter_rows()
            #first that correspond to the fields and data
            headers = [unicode(cell.value)
                       for cell in data_in_sheet.next()
                       ]
            headers = headers[1:]

            fields = [[cell.value for cell in row
                       if cell.value]
                      for row in data_in_sheet
                      ]
            headers_and_fields[sheet_name] = {'headers': headers,
                                              'fields': fields}
        return headers_and_fields

def test_by_type(cell, test_type):
    test = {'date': val.is_date, 'time': val.is_time, 'pulse': val.valid_pulse,'temp': val.valid_temp,
            'resp': val.valid_resp, 'systolic': val.valid_systolic, 'sodium': val.valid_sodium,
            'glucose': val.valid_glucose, 'hematocrit': val.valid_hematocrit, 'oxygen': val.valid_oxygen,
            'bun': val.valid_bun, 'blank': val.is_blank, 'ph': val.valid_ph
            }

    return test[test_type](cell)

def simple_check(field, tuple, statement):
    errors = []
    tuple = tuple
    simple_field_names = field[1:]
    for field_name in simple_field_names:
        error_found = eval("val.is_blank(" + statement % (field_name,) + ")")
        if error_found:
            errors.append(error_found)
    return errors


def complex_without_number_check(field, tuple, statement):
    errors = []
    tuple = tuple
    start_check = field[1]
    complex_field_names = field[2:]
    # check start condition
    if eval(statement % start_check) == True:
        for field_name in complex_field_names:
            error_found = eval("val.is_blank(" + statement % (field_name,) + ")")
            if error_found:
                errors.append(error_found)
    return errors

def complex_with_number_check(field, tuple, statement):
    errors = []
    tuple = tuple
    start_check = field[1]
    number_to_check = eval(statement % field[2])
    complex_field_names = field[3:]
    # check start condition
    if number_to_check:
        if eval(statement % start_check) == True:
            for i in range(1, number_to_check + 1):
                for field_name in complex_field_names:
                    error_found = eval("val.is_blank(" + statement % (field_name % i,) + ")")
                    if error_found:
                        errors.append(error_found)
    return errors

def simple_visit_check(field, tuple, statement, visit_num):
    errors = []
    if visit_num:
        tuple = tuple
        simple_field_names = field[1:]
        for field_name in simple_field_names:
            error_found = eval("val.is_blank(" + statement % (field_name % visit_num,) + ")")
            if error_found:
                errors.append(error_found)
    return errors

def complex_visit_with_number_check(field, tuple, statement, visit_num):
    errors = []
    if visit_num:
        tuple = tuple
        start_check = field[1] % visit_num
        number_to_check = eval(statement % field[2] % visit_num)
        if number_to_check:
            if number_to_check == "more than three":
                number_to_check = 3
            complex_field_names = field[3:]
            # check start condition
            if number_to_check >= 1:
                if eval(statement % start_check) == True:
                    for i in range(1, number_to_check + 1):
                        for field_name in complex_field_names:
                            error_found = eval("val.is_blank(" + statement % (field_name % (visit_num, i) + ")"))
                            if error_found:
                                errors.append(error_found)
    return errors


def complex_visit_without_number_check(field, tuple, statement, visit_num):
    errors = []
    if visit_num:
        tuple = tuple
        start_check = field[1] % visit_num
        complex_field_names = field[2:]
        # check start condition
        if eval(statement % start_check) == True:
            for field_name in complex_field_names:
                error_found = eval("val.is_blank(" + statement % (field_name % visit_num,) + ")")
                if error_found:
                    errors.append(error_found)
    return errors

def complex_visit_with_complex_number(field, tuple, statement, visit_num):
    errors = []
    if visit_num:
        tuple = tuple
        number_to_check = eval(statement % (field[1] % visit_num))
        if number_to_check:
            if number_to_check == 'more than three':
                number_to_check = 3
            if number_to_check >= 1:
                for i in range(1, number_to_check + 1):
                    start_check = field[2] % (visit_num, i)
                    complex_field_names = field[3:]
                    # check start condition
                    if eval(statement % start_check) == True:
                        for field_name in complex_field_names:
                            error_found = eval("val.is_blank(" + statement % (field_name % (visit_num, i) + ")"))
                            if error_found:
                                errors.append(error_found)
    return errors


def generic_check(check_type, file_to_check, headers, fields):
    print "startin %s check" % check_type
    with open(file_to_check,'rb') as datafile:
        data_workbook = openpyxl.load_workbook(datafile)
        data_worksheet = data_workbook.active
        data_fields = fields
        data_tuple = collections.namedtuple(check_type, field_names=headers)
        data = data_worksheet.iter_rows()
        #skip human readable headers in data file
        data.next()
        data_to_check = []
        for row in data:
            for_tuple = [cell for cell in row]
            one_subjects_data = data_tuple(*for_tuple)
            data_to_check.append(one_subjects_data)

        statement = "tuple.%s"
        all_found_errors = collections.defaultdict(dict)
        for tuple in data_to_check:
            errors = []
            # print tuple.id.value
            for field in data_fields:
                field_type = field[0]
                if field_type == 'simple:':
                    error_found = simple_check(field, tuple)
                    if error_found:
                        errors.append(error_found)
                if field_type == 'complex with number':
                    error_found = complex_with_number_check(field, tuple, statement)
                    if error_found:
                        errors.append(error_found)
                if field_type == 'complex without number':
                    error_found = complex_without_number_check(field, tuple, statement)
                    if error_found:
                        errors.append(error_found)
                if field_type == 'simple visit without number':
                    visit_num = tuple.visit_num.value
                    error_found = simple_visit_check(field, tuple, statement, visit_num)
                    if error_found:
                        errors.append(error_found)
                if field_type == 'complex visit with number':
                    visit_num = tuple.visit_num.value
                    error_found = complex_visit_with_number_check(field, tuple, statement, visit_num)
                    if error_found:
                        errors.append(error_found)
                if field_type == 'complex visit without number':
                    visit_num = tuple.visit_num.value
                    error_found = complex_visit_without_number_check(field, tuple, statement, visit_num)
                    if error_found:
                        errors.append(error_found)
                if field_type == 'complex visit with complex number':
                    visit_num = tuple.visit_num.value
                    error_found = complex_visit_with_complex_number(field, tuple, statement, visit_num)
                    if error_found:
                        errors.append(error_found)
            if errors != []:
                all_found_errors[tuple.id.value] = errors

        data_workbook.save(check_type + "_with_highlight_errors.xlsx")
        for key, value in all_found_errors.iteritems():
            print key
            for item in value:
                for err in item:
                    print err
            print "____________________________________________"


def eligibility_check():
    print "starting eligibility check"
    with open('14_0076_Eligibility.xlsx', 'rb') as eligibility_file:
        eligibility_workbook = openpyxl.load_workbook(eligibility_file)
        eligibility_worksheet = eligibility_workbook.active
        headers_and_fields = get_fields()
        eligibility_headers = headers_and_fields['eligibility']['headers']
        eligibility_fields = headers_and_fields['eligibility']['fields']

        eligibility_tuple = collections.namedtuple('eligibility', field_names=eligibility_headers)
        eligibility_data = eligibility_worksheet.iter_rows()
        #skip human readable headres eligibility_file
        eligibility_data.next()
        eligibility_data_to_check = []

        for row in eligibility_data:
            one_subjects_data = eligibility_tuple(*[cell
                                                    for cell in row
                                                    ])
            eligibility_data_to_check.append(one_subjects_data)

        statement = "eligibility_tuple.%s"
        for eligibility_tuple in eligibility_data_to_check:
            print eligibility_tuple.id.value
            for field in eligibility_fields:
                if field[0] == 'simple':
                    # print "simple ran"
                    simple_field_names = field[1:]
                    for field_name in simple_field_names:
                        eval("val.is_blank(" + statement % (field_name,) + ")")

                if field[0] == 'complex without number':
                    # print "complex without number ran"
                    start_check = field[1]
                    complex_field_names = field[2:]
                    #check start condition
                    if eval(statement % start_check) == True:
                        for field_name in complex_field_names:
                            eval("val.is_blank(" + statement % (field_name,) + ")")

                if field[0] == 'complex with number':
                    start_check = field[1]
                    number_to_check_cell = eval(field[2])
                    number_to_check = number_to_check_cell.value
                    complex_field_names = field[3:]
                    #check start condition
                    if eval(statement% start_check) == True:
                        for i in range(1, number_to_check + 1):
                            for field_name in complex_field_names:
                                eval("val.is_blank(" + statement % (field_name % i,) + ")")

        eligibility_workbook.save('14_0076_Eligibility_w_highlighted_erroros.xlsx')

def demographic_check():
    print "starting demographic check"
    with open('14_0076_Demographic.xlsx', 'rb') as demographic_file:
        demographic_workbook = openpyxl.load_workbook(demographic_file)
        demographic_worksheet = demographic_workbook.active
        headers_and_fields = get_fields()
        demographic_headers = headers_and_fields['demographic']['headers']
        demographic_fields = headers_and_fields['demographic']['fields']

        demographic_tuple = collections.namedtuple('demographic', field_names=demographic_headers)
        demographic_data = demographic_worksheet.iter_rows()
        #skip human readable headres demographic_file
        demographic_data.next()
        demographic_data_to_check = []

        for row in demographic_data:
            one_subjects_data = demographic_tuple(*[cell
                                                    for cell in row
                                                    ])
            demographic_data_to_check.append(one_subjects_data)

        statement = "demographic_tuple.%s"
        for demographic_tuple in demographic_data_to_check:
            print demographic_tuple.id.value
            for field in demographic_fields:
                if field[0] == 'simple':
                    # print "simple ran"
                    simple_field_names = field[1:]
                    for field_name in simple_field_names:
                        eval("val.is_blank(" + statement % (field_name,) + ")")

                if field[0] == 'complex without number':
                    # print "complex without number ran"
                    start_check = field[1]
                    complex_field_names = field[2:]
                    #check start condition
                    if eval(statement % start_check) == True:
                        for field_name in complex_field_names:
                            eval("val.is_blank(" + statement % (field_name,) + ")")

                if field[0] == 'complex with number':
                    start_check = field[1]
                    number_to_check_cell = eval(field[2])
                    number_to_check = number_to_check_cell.value
                    complex_field_names = field[3:]
                    #check start condition
                    if eval(statement % start_check) == True:
                        for i in range(1, number_to_check + 1):
                            for field_name in complex_field_names:
                                eval("val.is_blank(" + statement % (field_name % i,) + ")")

        demographic_workbook.save('14_0076_demographic_w_highlighted_erroros.xlsx')

def symptoms_check():
    print "starting symptoms check"
    with open('14_0076_Symptoms.xlsx', 'rb') as symptoms_file:
        symptoms_workbook = openpyxl.load_workbook(symptoms_file)
        symptoms_worksheet = symptoms_workbook.active
        headers_and_fields = get_fields()
        symptoms_headers = headers_and_fields['symptoms']['headers']
        symptoms_fields = headers_and_fields['symptoms']['fields']

        symptoms_tuple = collections.namedtuple('symptoms', field_names=symptoms_headers)
        symptoms_data = symptoms_worksheet.iter_rows()
        #skip human readable headres symptoms_file
        symptoms_data.next()
        symptoms_data_to_check = []

        for row in symptoms_data:
            one_subjects_data = symptoms_tuple(*[cell
                                                    for cell in row
                                                    ])
            symptoms_data_to_check.append(one_subjects_data)

        statement = "symptoms_tuple.%s"
        for symptoms_tuple in symptoms_data_to_check:
            print symptoms_tuple.id.value
            for field in symptoms_fields:
                if field[0] == 'simple':
                    # print "simple ran"
                    simple_field_names = field[1:]
                    for field_name in simple_field_names:
                        eval("val.is_blank(" + statement % (field_name,) + ")")

                if field[0] == 'complex without number':
                    # print "complex without number ran"
                    start_check = field[1]
                    complex_field_names = field[2:]
                    #check start condition
                    if eval("symptoms_tuple.%s" % start_check) == True:
                        for field_name in complex_field_names:
                            eval("val.is_blank(" + statement % (field_name,) + ")")

                if field[0] == 'complex with number':
                    start_check = field[1]
                    number_to_check_cell = eval(field[2])
                    number_to_check = number_to_check_cell.value
                    complex_field_names = field[3:]
                    #check start condition
                    if eval("symptoms_tuple.%s" % start_check) == True:
                        for i in range(1, number_to_check + 1):
                            for field_name in complex_field_names:
                                eval("val.is_blank(" + statement % (field_name % i,) + ")")

        symptoms_workbook.save('14_0076_symptoms_w_highlighted_erroros.xlsx')

def medical_check():
    print "starting medical check"
    with open('14_0076_medical.xlsx', 'rb') as medical_file:
        medical_workbook = openpyxl.load_workbook(medical_file)
        medical_worksheet = medical_workbook.active
        headers_and_fields = get_fields()
        medical_headers = headers_and_fields['medical']['headers']
        medical_fields = headers_and_fields['medical']['fields']

        medical_tuple = collections.namedtuple('medical', field_names=medical_headers)
        medical_data = medical_worksheet.iter_rows()
        #skip human readable headres medical_file
        medical_data.next()
        medical_data_to_check = []

        for row in medical_data:
            one_subjects_data = medical_tuple(*[cell
                                                    for cell in row
                                                    ])
            medical_data_to_check.append(one_subjects_data)

        statement = "medical_tuple.%s"
        for medical_tuple in medical_data_to_check:
            print medical_tuple.id.value
            for field in medical_fields:
                if field[0] == 'simple':
                    # print "simple ran"
                    simple_field_names = field[1:]
                    for field_name in simple_field_names:
                        eval("val.is_blank(" + statement % (field_name,) + ")")

                if field[0] == 'complex without number':
                    # print "complex without number ran"
                    start_check = field[1]
                    complex_field_names = field[2:]
                    #check start condition
                    if eval(statement % start_check) == True:
                        for field_name in complex_field_names:
                            eval("val.is_blank(" + statement % (field_name,) + ")")

                if field[0] == 'complex with number':
                    start_check = field[1]
                    number_to_check_cell = eval(statement % field[2])
                    number_to_check = number_to_check_cell.value
                    complex_field_names = field[3:]
                    #check start condition
                    if eval(statement% start_check) == True:
                        for i in range(1, number_to_check + 1):
                            for field_name in complex_field_names:
                                eval("val.is_blank(" + statement % (field_name % i,) + ")")

        medical_workbook.save('14_0076_medical_w_highlighted_erroros.xlsx')

def enrollment_specimen_check():
    print "starting enrollment specimen check"
    with open('14_0076_Enrollment_Specimen.xlsx', 'rb') as enrollment_specimen_file:
        enrollment_specimen_workbook = openpyxl.load_workbook(enrollment_specimen_file)
        enrollment_specimen_worksheet = enrollment_specimen_workbook.active
        headers_and_fields = get_fields()
        enrollment_specimen_headers = headers_and_fields['enrollment_specimen']['headers']
        enrollment_specimen_fields = headers_and_fields['enrollment_specimen']['fields']

        enrollment_specimen_tuple = collections.namedtuple('enrollment_specimen', field_names=enrollment_specimen_headers)
        enrollment_specimen_data = enrollment_specimen_worksheet.iter_rows()
        #skip human readable headres enrollment_specimen_file
        enrollment_specimen_data.next()
        enrollment_specimen_data_to_check = []

        for row in enrollment_specimen_data:
            one_subjects_data = enrollment_specimen_tuple(*[cell
                                                    for cell in row
                                                    ])
            enrollment_specimen_data_to_check.append(one_subjects_data)

        statement = "enrollment_specimen_tuple.%s"
        for enrollment_specimen_tuple in enrollment_specimen_data_to_check:
            print enrollment_specimen_tuple.id.value
            for field in enrollment_specimen_fields:
                if field[0] == 'simple':
                    # print "simple ran"
                    simple_field_names = field[1:]
                    for field_name in simple_field_names:
                        eval("val.is_blank(" + statement % (field_name,) + ")")

                if field[0] == 'complex without number':
                    # print "complex without number ran"
                    start_check = field[1]
                    complex_field_names = field[2:]
                    #check start condition
                    if eval(statement % start_check) == True:
                        for field_name in complex_field_names:
                            eval("val.is_blank(" + statement % (field_name,) + ")")

                if field[0] == 'complex with number':
                    start_check = field[1]
                    number_to_check_cell = eval(field[2])
                    number_to_check = number_to_check_cell.value
                    complex_field_names = field[3:]
                    #check start condition
                    if eval(statement% start_check) == True:
                        for i in range(1, number_to_check + 1):
                            for field_name in complex_field_names:
                                eval("val.is_blank(" + statement % (field_name % i,) + ")")

        enrollment_specimen_workbook.save('14_0076_enrollment_specimen_w_highlighted_erroros.xlsx')

def follow_up_assessment_check():
    print "starting follow_up assessment check"
    with open('14_0076_Follow_Up_Assessment.xlsx', 'rb') as follow_up_assessment_file:
        follow_up_assessment_workbook = openpyxl.load_workbook(follow_up_assessment_file)
        follow_up_assessment_worksheet = follow_up_assessment_workbook.active
        headers_and_fields = get_fields()
        follow_up_assessment_headers = headers_and_fields['follow_up_assessment']['headers']
        follow_up_assessment_fields = headers_and_fields['follow_up_assessment']['fields']

        follow_up_assessment_tuple = collections.namedtuple('follow_up_assessment', field_names=follow_up_assessment_headers)
        follow_up_assessment_data = follow_up_assessment_worksheet.iter_rows()
        #skip human readable headres follow_up_assessment_file
        follow_up_assessment_data.next()
        follow_up_assessment_data_to_check = []

        for row in follow_up_assessment_data:
            one_subjects_data = follow_up_assessment_tuple(*[cell
                                                    for cell in row
                                                    ])
            follow_up_assessment_data_to_check.append(one_subjects_data)

        statement = "follow_up_assessment_tuple.%s"
        for follow_up_assessment_tuple in follow_up_assessment_data_to_check:
            print follow_up_assessment_tuple.id.value
            for field in follow_up_assessment_fields:
                if field[0] == 'simple':
                    # print "simple ran"
                    simple_field_names = field[1:]
                    for field_name in simple_field_names:
                        eval("val.is_blank(" + statement % (field_name,) + ")")

                if field[0] == 'complex without number':
                    # print "complex without number ran"
                    start_check = field[1]
                    complex_field_names = field[2:]
                    #check start condition
                    if eval(statement % start_check) == True:
                        for field_name in complex_field_names:
                            eval("val.is_blank(" + statement % (field_name,) + ")")

                if field[0] == 'complex with number':
                    start_check = field[1]
                    number_to_check_cell = eval(statement % field[2])
                    number_to_check = number_to_check_cell.value
                    complex_field_names = field[3:]
                    #check start condition
                    if eval(statement % start_check) == True:
                        for i in range(1, number_to_check + 1):
                            for field_name in complex_field_names:
                                eval("val.is_blank(" + statement % (field_name % i,) + ")")

        follow_up_assessment_workbook.save('14_0076_follow_up_assessment_w_highlighted_erroros.xlsx')


def follow_up_specimen_check():
    print "starting follow_up specimen check"
    with open('14_0076_Follow_Up_Specimen.xlsx', 'rb') as follow_up_specimen_file:
        follow_up_specimen_workbook = openpyxl.load_workbook(follow_up_specimen_file)
        follow_up_specimen_worksheet = follow_up_specimen_workbook.active
        headers_and_fields = get_fields()
        follow_up_specimen_headers = headers_and_fields['follow_up_specimen']['headers']
        follow_up_specimen_fields = headers_and_fields['follow_up_specimen']['fields']

        follow_up_specimen_tuple = collections.namedtuple('follow_up_specimen',
                                                            field_names=follow_up_specimen_headers)
        follow_up_specimen_data = follow_up_specimen_worksheet.iter_rows()
        # skip human readable headres follow_up_specimen_file
        follow_up_specimen_data.next()
        follow_up_specimen_data_to_check = []

        for row in follow_up_specimen_data:
            one_subjects_data = follow_up_specimen_tuple(*[cell
                                                             for cell in row
                                                             ])
            follow_up_specimen_data_to_check.append(one_subjects_data)

        statement = "follow_up_specimen_tuple.%s"
        for follow_up_specimen_tuple in follow_up_specimen_data_to_check:
            print follow_up_specimen_tuple.id.value
            for field in follow_up_specimen_fields:
                if field[0] == 'simple':
                    # print "simple ran"
                    simple_field_names = field[1:]
                    for field_name in simple_field_names:
                        eval("val.is_blank(" + statement % (field_name,) + ")")

                if field[0] == 'complex without number':
                    # print "complex without number ran"
                    start_check = field[1]
                    complex_field_names = field[2:]
                    # check start condition
                    if eval(statement % start_check) == True:
                        for field_name in complex_field_names:
                            eval("val.is_blank(" + statement % (field_name,) + ")")

                if field[0] == 'complex with number':
                    start_check = field[1]
                    number_to_check_cell = statement % field[2]
                    number_to_check = number_to_check_cell.value
                    complex_field_names = field[3:]
                    # check start condition
                    if eval(statement % start_check) == True:
                        for i in range(1, number_to_check + 1):
                            for field_name in complex_field_names:
                                eval("val.is_blank(" + statement % (field_name % i,) + ")")

        follow_up_specimen_workbook.save('14_0076_follow_up_specimen_w_highlighted_erroros.xlsx')


def ed_visit_check():
    print "start ed visit check"
    with open('14_0076_Ed_Visit.xlsx', 'rb') as ed_visit_file:
        ed_visit_workbook = openpyxl.load_workbook(ed_visit_file)
        ed_visit_worksheet = ed_visit_workbook.active
        headers_and_fields = get_fields()
        ed_visit_headers = headers_and_fields['ed_visit']['headers']
        ed_visit_fields = headers_and_fields['ed_visit']['fields']

        ed_visit_tuple = collections.namedtuple('ed_visit',
                                                field_names=ed_visit_headers)
        ed_visit_data = ed_visit_worksheet.iter_rows()
        # skip human readable headres ed_visit_file
        ed_visit_data.next()
        ed_visit_data_to_check = []

        for row in ed_visit_data:
            one_subjects_data = ed_visit_tuple(*[cell
                                                 for cell in row
                                                 ])
            ed_visit_data_to_check.append(one_subjects_data)

        statement = "ed_visit_tuple.%s"
        for ed_visit_tuple in ed_visit_data_to_check:
            print ed_visit_tuple.id.value
            # check for ed visit number
            ed_visit_num = ed_visit_tuple.visit_num.value
            if ed_visit_num >= 1:
                for i in range(1, ed_visit_num + 1):
                    for field in ed_visit_fields:
                        if field[0] == 'simple':
                            # print "simple ran"
                            simple_field_names = field[1:]
                            for field_name in simple_field_names:
                                eval("val.is_blank(" + statement % (field_name,) + ")")

                        if field[0] == 'simple visit without number':
                            # print "simple visit ran"
                            simple_field_names = field[1:]
                            for field_name in simple_field_names:
                                eval("val.is_blank(" + statement % (field_name % ed_visit_num,) + ")")

                        if field[0] == 'complex visit without number':
                            # print "complex visit without number ran"
                            start_check = field[1] % ed_visit_num
                            complex_field_names = field[2:]
                            # check start condition
                            if eval(statement % start_check) == True:
                                for field_name in complex_field_names:
                                    eval("val.is_blank(" + statement % (field_name % ed_visit_num,) + ")")

                        if field[0] == 'complex visit with number':
                            # print "complex visit with number ran"
                            start_check = field[1] % ed_visit_num
                            number_to_check_cell = eval(statement % field[2] % ed_visit_num)
                            number_to_check = number_to_check_cell.value
                            if number_to_check == "more than three":
                                number_to_check = 3
                            complex_field_names = field[3:]
                            # check start condition
                            if number_to_check >= 1:
                                if eval(statement % start_check) == True:
                                    for i in range(1, number_to_check + 1):
                                        for field_name in complex_field_names:
                                            eval("val.is_blank(" + statement % (field_name % (ed_visit_num, i) + ")"))

                        if field[0] == 'complex visit with complex number':
                            # print "complex visit with complex number ran"
                            number_to_check_cell = eval(statement % (field[1] % ed_visit_num))
                            number_to_check = number_to_check_cell.value
                            if number_to_check == 'more than three':
                                number_to_check = 3
                            if number_to_check >= 1:
                                for i in range(1, number_to_check + 1):
                                    start_check = field[2] % (ed_visit_num, i)
                                    complex_field_names = field[3:]
                                    # check start condition
                                    if eval(statement % start_check) == True:
                                        for field_name in complex_field_names:
                                            eval("val.is_blank(" + statement % (field_name % (ed_visit_num,i) + ")"))

        ed_visit_workbook.save('14_0076_ed_visit_w_highlighted_erroros.xlsx')


def ip_visit_check():
    print "start IP visit check"
    with open('14_0076_IP_Visit.xlsx', 'rb') as ip_visit_file:
        ip_visit_workbook = openpyxl.load_workbook(ip_visit_file)
        ip_visit_worksheet = ip_visit_workbook.active
        headers_and_fields = get_fields()
        ip_visit_headers = headers_and_fields['ip_visit']['headers']
        ip_visit_fields = headers_and_fields['ip_visit']['fields']

        ip_visit_tuple = collections.namedtuple('ip_visit',
                                                field_names=ip_visit_headers)
        ip_visit_data = ip_visit_worksheet.iter_rows()
        # skip human readable headres ip_visit_file
        ip_visit_data.next()
        ip_visit_data_to_check = []

        for row in ip_visit_data:
            one_subjects_data = ip_visit_tuple(*[cell
                                                 for cell in row
                                                 ])
            ip_visit_data_to_check.append(one_subjects_data)

        statement = "ip_visit_tuple.%s"
        for ip_visit_tuple in ip_visit_data_to_check:
            print ip_visit_tuple.id.value
            # check for ed visit number
            ip_visit_num = ip_visit_tuple.visit_num.value
            if ip_visit_num >= 1:
                for i in range(1, ip_visit_num + 1):
                    for field in ip_visit_fields:
                        if field[0] == 'simple':
                            # print "simple ran"
                            simple_field_names = field[1:]
                            for field_name in simple_field_names:
                                eval("val.is_blank(" + statement % (field_name,) + ")")

                        if field[0] == 'simple visit without number':
                            # print "simple visit ran"
                            simple_field_names = field[1:]
                            for field_name in simple_field_names:
                                eval("val.is_blank(" + statement % (field_name % ip_visit_num,) + ")")

                        if field[0] == 'complex visit without number':
                            # print "complex visit without number ran"
                            start_check = field[1] % ip_visit_num
                            complex_field_names = field[2:]
                            # check start condition
                            if eval(statement % start_check) == True:
                                for field_name in complex_field_names:
                                    eval("val.is_blank(" + statement % (field_name % ip_visit_num,) + ")")

                        if field[0] == 'complex visit with number':
                            # print "complex visit with number ran"
                            start_check = field[1] % ip_visit_num
                            number_to_check_cell = eval(statement % field[2] % ip_visit_num)
                            number_to_check = number_to_check_cell.value
                            if number_to_check == "more than three":
                                number_to_check = 3
                            complex_field_names = field[3:]
                            # check start condition
                            if number_to_check >= 1:
                                if eval(statement % start_check) == True:
                                    for i in range(1, number_to_check + 1):
                                        for field_name in complex_field_names:
                                            eval("val.is_blank(" + statement % (field_name % (ip_visit_num, i) + ")"))

                        if field[0] == 'complex visit with complex number':
                            # print "complex visit with complex number ran"
                            number_to_check_cell = eval(statement % (field[1] % ip_visit_num))
                            number_to_check = number_to_check_cell.value
                            if number_to_check == 'more than three':
                                number_to_check = 3
                            if number_to_check >= 1:
                                for i in range(1, number_to_check + 1):
                                    start_check = field[2] % (ip_visit_num, i)
                                    complex_field_names = field[3:]
                                    # check start condition
                                    if eval(statement % start_check) == True:
                                        for field_name in complex_field_names:
                                            eval("val.is_blank(" + statement % (field_name % (ip_visit_num,i) + ")"))

        ip_visit_workbook.save('14_0076_ip_visit_w_highlighted_erroros.xlsx')


def main():
    fields = get_fields()
    for check_type, headers_and_fields in fields.iteritems():
        filename = "14_0076_{}.xlsx".format(check_type)
        print filename
        headers = headers_and_fields['headers']
        fields = headers_and_fields['fields']
        # print fields
        generic_check(check_type,filename,headers,fields)

if __name__ == '__main__':
    main()